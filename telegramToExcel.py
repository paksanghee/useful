import telegram
from openpyxl import load_workbook
import openpyxl
import xlsxwriter
#https://api.telegram.org/bot1661487143:AAHOSggWOMG2bth3TrBkEGQ6FFwyfDgWbm4/getUpdates

#rb=load_workbook("")
#ws = rb['Sheet1']
#print(ws['C4'].value)
fileName = ''
Storage_path = ''

ExelList=[] #엑셀에 들어갈 내용만 
TOKEN = ''
CHAT_ID = ''


# 특정 값 찾는 함수
def find_index(data, target):
  res = []
  lis = data
  while True:
    try:
      res.append(lis.index(target) + (res[-1]+1 if len(res)!=0 else 0)) #+1의 이유 : 0부터 시작이니까
      lis = data[res[-1]+1:]
    except:
      break     
  return res

'''
def IfNewMakeNewFile(mylist):
    global fileName
    for i in mylist:
        if '신규:' in i:           
            new=i.split(':')
            fileName = new[1] + ".xlsx"
            create = xlsxwriter.Workbook(Storage_path+fileName)           
            #create.save(Storage_path+fileName)
            create.close()
            print(fileName+" file created")
'''

def FindFileNameAndGet(mylist):
    global fileName
    for i in mylist:
        if 'file:' in i:           
            old=i.split(':')
            fileName = old[1] + ".xlsx"            
            print(fileName+" file writing..")

def LeaveOnlyLinesContainingOnlyCertainChar(mylist):
    global ExelList
    CertainCharacters = '/' # '/'를 포함하는 줄만 남기기
    for i in mylist: 
        if CertainCharacters in i and 'ex' not in i :
            ExelList.append(i) 
    print("leave only function end")        


def DividedByCharSavedInExcel(sheet):
    global ExelList
    splitCellBy = '/'
    for i in ExelList: 
        c=i.split(splitCellBy)
        sheet.append(c)   
    print("DividedByChar and SavedInExcel..!")






bot = telegram.Bot(token=TOKEN)
#bot.sendMessage(chat_id = CHAT_ID, text = '하이') #
updates = bot.getUpdates()  #해당 봇의 업데이트 내역 받아옴 (사용자의 메세지 등)
for u in reversed(updates) :  # 최근 것 하나만 출력
    telemsg=u.message.text
    break

arr=telemsg.split('\n')

FindFileNameAndGet(arr) #파일이름 추출


#엑셀
wb = openpyxl.Workbook()
sheet1 = wb.active
sheet2 = wb.create_sheet('result')
sheet2.title = 'example'



'''
while '' in arr: #빈 줄 제거
    arr.remove('')

# ㅡㅡㅡㅡㅡ 사이만 남기기
index=find_index(arr,'ㅡㅡㅡㅡㅡ')
exelList=arr[index[0]+1:index[1]]
'''
LeaveOnlyLinesContainingOnlyCertainChar(arr)
DividedByCharSavedInExcel(sheet1)    
  

'''

for i in arr:
    if i == "ㅡㅡㅡㅡㅡ":
        break    
    else:
        del arr[j]
        j=j+1

j=0
for i in arr:
    if i == '':
        pass
    else:
        j=j+1
        print(i+"...."+str(j))
'''

 
wb.save(Storage_path+fileName)
